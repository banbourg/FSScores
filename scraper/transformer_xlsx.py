# -*- coding: utf-8 -*-
# #!/bin/env python


import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

import os
import re
import sys
import logging

logging.basicConfig(#filename="transformer" + datetime.today().strftime("%Y-%m-%d_%H-%M-%S") + ".log",
                    format="%(asctime)s - %(name)s - %(levelname)-5s - %(message)s",
                    level=logging.INFO,
                    datefmt="%Y-%m-%d %H:%M:%S")

logger = logging.getLogger(__name__)

p_list = [os.path.abspath("./classes/"), os.path.abspath("..")]
for p in p_list:
    if p not in sys.path:
        sys.path.append(p)

try:
    import settings
    import event
    import person
    import protocol
    import db_builder
except ImportError as exc:
    sys.exit(f"Error: failed to import module ({exc})")


# ------------------------------------------ CHANGE RUN PARAMETERS HERE ---------------------------------------------
ENABLE_PAUSE = False
# ----------------------------------------------------------------------------------------------------------------------

ABBREV_DIC = {'gpjpn': 'NHK', 'gpfra': 'TDF', 'gpcan': 'SC', 'gprus': 'COR', 'gpusa': 'SA', 'gpchn': 'COC',
             'gpf': 'GPF', 'wc': 'WC', 'fc': '4CC', 'owg': 'OWG', 'wtt': 'WTT', 'sc': 'SC', 'ec': 'EC', 'sa': 'SA',
             'jgpfra': 'JGPFRA', 'nhk': 'NHK'}


def reorder_cols(df, col_to_move, new_pos):
    cols = df.columns.tolist()
    cols.insert(new_pos, cols.pop(cols.index(col_to_move)))
    return df[cols]


def find_protocol_coordinates(df):
    protocol_starts, protocol_ends = [], []
    for i in df.index:
        for j in range(0,6):
            if "Name" in str(df.iloc[i, j]):
                protocol_starts.append(i)
            if "Deductions" in str(df.iloc[i, j]) and j < 4:
                protocol_ends.append(i)
    return list(zip(protocol_starts, protocol_ends))


def scrape_sheet(df, segment, last_row_dic, skater_list, conn_dic):
    protocol_coords = find_protocol_coordinates(df)
    logger.debug(f"Protocol coordinates are {protocol_coords}")

    for c in protocol_coords:
        prot = protocol.Protocol(df=df,
                                 protocol_coordinates=c,
                                 segment=segment,
                                 last_row_dic=last_row_dic,
                                 skater_list=skater_list,
                                 conn_dic=conn_dic)
        for i in prot.row_range:
            for j in prot.col_range:
                if "Skating Skills" in str(df.iloc[i, j]):
                    try:
                        prot.parse_pcs_table(df, i, j, last_row_dic)
                    except ValueError as ve:
                        logger.error(f"Encountered error reading PCS row in {segment.name} {segment.year} "
                                     f"{segment.discipline} {segment.segment}, {dict(vars(prot.skater))}: {ve}")
                elif "Elements" in str(df.iloc[i, j]):
                    try:
                        prot.parse_tes_table(df, i, j, last_row_dic)
                    except ValueError as ve:
                        logger.error(f"Encountered error reading TES row in {segment.name} {segment.year} "
                                     f"{segment.discipline} {segment.segment}, {dict(vars(prot.skater))}: {ve}")
                elif "Deductions" in str(df.iloc[i, j]) and j < 4:
                    prot.parse_deductions(df, i, j, segment)
        segment.protocol_list.append(prot)


def convert_to_dfs(segment_list, conn_dic, competitor_list, id_dic):
    all_dics = {"segments": [], "competitors": [], "protocols": [], "pcs_averages": [], "pcs_detail": [],
                "deductions_detail": [], "elements": [], "goe_detail": []}
    all_dfs = {}

    for s in segment_list:
        if "segments" in all_dics:
            all_dics["segments"].append(s.get_segment_dic())

        for p in s.protocol_list:
            all_dics["protocols"].append(p.get_protocol_dic(segment=s))

            all_dics["pcs_averages"].extend(p.pcs_av_list)
            all_dics["pcs_detail"].extend(p.pcs_detail_list)

            all_dics["deductions_detail"].append(p.get_deductions_dic())

            for e in p.elts:
                elt_dic = e.get_element_dic()
                elt_dic["protocol_id"] = p.id
                all_dics["elements"].append(elt_dic)

                if e.goe_dic:
                    all_dics["goe_detail"].append(e.goe_dic)

    for c in competitor_list:
        all_dics["competitors"].append(c.get_competitor_dict())

    for key in all_dics:
        all_dfs[key] = pd.DataFrame(all_dics[key])

    # panels_df = pd.read_sql_query("SELECT * FROM panels", conn_dic["engine"])
    dic = {"pcs": ["pcs_avg", "judge_no"],
           "goe": ["element", "judge_no"],
           "deductions": ["protocol", "deduction_type"]}

    for s in dic:
        key = s + "_detail"
        crossref_id = dic[s][0] + "_id"
        all_dfs[key] = all_dfs[key].melt(id_vars=[crossref_id], var_name=dic[s][1], value_name=s + "_score")
        all_dfs[key] = all_dfs[key][all_dfs[key][s + "_score"].notnull()]
        all_dfs[key].insert(0, "id", range(id_dic[key], id_dic[key] + len(all_dfs[key])))
        id_dic[key] += (len(all_dfs[key]) + 1)
    #     all_dfs[key] = pd.merge(all_dfs[key], panels_df, how='left', left_on="judge_no", right_on=["official_role"])

    return all_dfs


def transform_and_load(read_path, write_path, naming_schema, counter, db_credentials):
    done_dir_path = os.path.join(read_path, "done")
    if not os.path.exists(done_dir_path):
        os.makedirs(done_dir_path)

    # --- 1. Initiate connections and fetch ids
    conn, engine = db_builder.initiate_connections(db_credentials)
    cur = conn.cursor()
    conn_dic = {"conn": conn, "engine": engine, "cursor": cur}

    # --- 2. Get max table rows for append
    rows = {}
    for x in ["deductions_detail", "pcs_averages", "pcs_detail", "goe_detail", "elements", "segments", "competitors",
              "officials", "panels", "protocols"]:
        rows[x] = db_builder.get_last_row_key(table_name=x, cursor=cur) + 1

    # --- 3. Iteratively read through converted .xlsx and populate tables
    segment_list, skater_list = [], []
    file_count = 0
    files = sorted(glob.glob(read_path + '*.xlsx'))

    for f in files:
        filename = f.rpartition("/")[2]
        basename = filename.rpartition(".")[0]
        logger.info(f"Attempting to read {basename}")
        file_count += 1

        try:
            disc = event.parse_discipline(filename)
        except ValueError:
            logger.error(f"Passing on file {filename}")
            continue

        try:
            seg = event.ScoredSegment(name_to_parse=basename, discipline=disc, id_dic=rows)
        except ValueError as ve:
            logger.error(f"Failed to instantiate ScoredSegment object: {ve}")

        segment_list.append(seg)

        wb = load_workbook(f)
        for sheet in wb.sheetnames:
            raw_df = pd.DataFrame(wb[sheet].values)
            scrape_sheet(df=raw_df, segment=seg, last_row_dic=rows, skater_list=skater_list, conn_dic=conn_dic)

        if file_count % counter == 0:
            dfs = convert_to_dfs(segment_list=segment_list, competitor_list=skater_list, conn_dic=conn_dic, id_dic=rows)

            for k in dfs:
                db_builder.create_staging_table(df=dfs[k], conn_dic=conn_dic, table_name=k, fetch_last_row=False)

            if ENABLE_PAUSE:
                input("Hit Enter to write to main tables")

            for k in dfs:
                db_builder.write_to_final_table(df=dfs[k], conn_dic=conn_dic, table_name=k)
                write_to_csv(df=dfs[k], table_name=k, write_path=write_path)

            segment_list, skater_list = [], []

        current_path = os.path.join(read_path, filename)
        done_path = os.path.join(done_dir_path, filename)
        os.rename(current_path, done_path)


def write_to_csv(df, table_name, write_path):
    file_path = os.path.join(write_path, table_name + ".csv")
    if os.path.isfile(file_path):
        include_header = False
    else:
        include_header = True
    df.to_csv(header=include_header, index=False, mode="a", date_format="%y-%m-%d")



def main():
    # # FOR YOGEETA AND YOGEETA ONLY
    # try:
    #     assert len(sys.argv) in [3,4]
    # except AssertionError:
    #     sys.exit("Please pass in the path to the xlsx dir, the path to which the csv backup should be written, and
    #               (optionally) an integer counter (for frequency at which results get written to the staging table,
    #               default is every 10 files), in that order")
    #
    # read_path = sys.argv[1]
    # write_path = sys.argv[2]
    # try:
    #     counter = int(sys.argv[1])
    # except (IndexError, TypeError):
    #     counter = 10

    db_credentials = settings.DB_CREDENTIALS
    read_path = settings.READ_PATH
    write_path = settings.WRITE_PATH
    counter = 1
    naming_schema = "_new"

    transform_and_load(read_path, write_path, naming_schema, counter, db_credentials)


if __name__ == "__main__":
    main()